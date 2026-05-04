from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from .models import *
from .models import Empleado, Bitacora
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from .models import CargaDescarga
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime


def registrar_accion(empleado, accion, descripcion, request=None):
    """
    Función auxiliar para registrar acciones en la bitácora
    """
    ip_address = None
    if request:
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip_address = x_forwarded_for.split(',')[0]
        else:
            ip_address = request.META.get('REMOTE_ADDR')
    
    Bitacora.objects.create(
        empleado=empleado,
        accion=accion,
        descripcion=descripcion,
        ip_address=ip_address
    )


def iniciar_sesion(request):
    if request.method == 'POST':
        nombre = request.POST['nombre']
        clave_empleado = request.POST['clave_empleado']
        
        try:
            # Buscar al empleado por su clave de empleado y nombre
            empleado = Empleado.objects.get(clave_empleado=clave_empleado, nombre=nombre)
            
            # Guardar el ID del empleado en la sesión
            request.session['empleado_id'] = empleado.id
            
            # Registrar en bitácora
            registrar_accion(empleado, 'Inicio de sesión', f'El usuario {empleado.nombre} inició sesión', request)
            
            return redirect('bienvenida')
        except Empleado.DoesNotExist:
            # Mostrar mensaje de error si las credenciales son incorrectas
            return render(request, 'iniciar_sesion.html', {'error': 'Credenciales incorrectas'})
    
    return render(request, 'iniciar_sesion.html')

def bienvenida(request):
    # Verificar si el empleado está autenticado
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    # Obtener el empleado autenticado
    empleado = Empleado.objects.get(id=empleado_id)
    
    # Obtener registros de la tabla expedientes
    expedientes = ConciliacionExpedientes.objects.all()
    
    return render(request, 'bienvenida.html', {
        'empleado': empleado,
        'expedientes': expedientes,
    })


def cerrar_sesion(request):
    # Registrar en bitácora antes de cerrar sesión
    empleado_id = request.session.get('empleado_id')
    if empleado_id:
        try:
            empleado = Empleado.objects.get(id=empleado_id)
            registrar_accion(empleado, 'Cierre de sesión', f'El usuario {empleado.nombre} cerró sesión', request)
        except Empleado.DoesNotExist:
            pass
    
    # Eliminar el ID del empleado de la sesión
    if 'empleado_id' in request.session:
        del request.session['empleado_id']
    return redirect('iniciar_sesion')

def agregar_expediente(request):
    # Verificar si el usuario está autenticado
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=empleado_id)
    
    if request.method == 'POST':
        # Get the last id_expediente and increment it
        last_expediente = ConciliacionExpedientes.objects.order_by('-id_expediente').first()
        new_id = 1 if not last_expediente else last_expediente.id_expediente + 1
        
        nuevo_expediente = ConciliacionExpedientes(
            id_expediente=new_id,
            expediente=request.POST['expediente'],
            junta=request.POST.get('junta', None),
            tomo=request.POST.get('tomo', None),
            actor_nombre=request.POST.get('actor_nombre', None),
            mujer_numero=request.POST.get('mujer_numero', None),
            hombre_numero=request.POST.get('hombre_numero', None),
            demandado_nombre=request.POST.get('demandado_nombre', None),
            mujer_numero_0=request.POST.get('mujer_numero_0', None),
            hombre_numero_0=request.POST.get('hombre_numero_0', None),
            persona_moral_numero=request.POST.get('persona_moral_numero', None),
            iebem=request.POST.get('iebem', None),
            servicios_salud=request.POST.get('servicios_salud', None),
            poder_ejecutivo=request.POST.get('poder_ejecutivo', None),
            ayuntamientos=request.POST.get('ayuntamientos', None),
            otros_organismos=request.POST.get('otros_organismos', None),
            no_se_ha_notificado_a_las_partes=request.POST.get('no_se_ha_notificado_a_las_partes', None),
            empl_no_realizado=request.POST.get('empl_no_realizado', None),
            empl_exhorto=request.POST.get('empl_exhorto', None),
            exhortos_sin_enviar=request.POST.get('exhortos_sin_enviar', None),
            exhortos_cdmx=request.POST.get('exhortos_cdmx', None),
            exhortos_foraneos=request.POST.get('exhortos_foraneos', None),
            cde=request.POST.get('cde', None),
            tercero_llamado_a_juicio=request.POST.get('tercero_llamado_a_juicio', None),
            oap=request.POST.get('oap', None),
            desahogo_pruebas=request.POST.get('desahogo_pruebas', None),
            test_falta_citar=request.POST.get('test_falta_citar', None),
            periciales_partes=request.POST.get('periciales_partes', None),
            pericial_tercero=request.POST.get('pericial_tercero', None),
            inf_falta_hacer=request.POST.get('inf_falta_hacer', None),
            inf_falta_desahogar=request.POST.get('inf_falta_desahogar', None),
            otras_pruebas=request.POST.get('otras_pruebas', None),
            alegatos=request.POST.get('alegatos', None),
            prueba_pendiente=request.POST.get('prueba_pendiente', None),
            cierre=request.POST.get('cierre', None),
            laudo_dictado=request.POST.get('laudo_dictado', None),
            absolutorio=request.POST.get('absolutorio', None),
            condenatorio=request.POST.get('condenatorio', None),
            monto=request.POST.get('monto', None),
            auto_ejecucion=request.POST.get('auto_ejecucion', None),
            terceria=request.POST.get('terceria', None),
            recurso_revision=request.POST.get('recurso_revision', None),
            remate=request.POST.get('remate', None),
            inactividad_1_anio=request.POST.get('inactividad_1_anio', None),
            inactividad_2_anios_mas=request.POST.get('inactividad_2_anios_mas', None),
            amparo_indirecto=request.POST.get('amparo_indirecto', None),
            amparo_directo_cumpl=request.POST.get('amparo_directo_cumpl', None),
            sustitucion_patronal=request.POST.get('sustitucion_patronal', None),
            no_interpuesta=request.POST.get('no_interpuesta', None),
            prescripcion=request.POST.get('prescripcion', None),
            depuracion=request.POST.get('depuracion', None),
            regularizar=request.POST.get('regularizar', None),
            reviso_capturo=request.POST.get('reviso_capturo', None)
        )
        nuevo_expediente.save()
        
        # Registrar en bitácora
        registrar_accion(empleado, 'Crear expediente', f'Creó el expediente {request.POST["expediente"]} - Actor: {request.POST.get("actor_nombre", "")} vs Demandado: {request.POST.get("demandado_nombre", "")}', request)
        
        messages.success(request, 'Expediente agregado correctamente.')
        return redirect('bienvenida')
    
    return render(request, 'agregar_expediente.html')


def crear_empleado(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('bienvenida')

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        clave_empleado = request.POST.get('clave_empleado')
        puesto = request.POST.get('puesto')
        if nombre and clave_empleado and puesto:
            if Empleado.objects.filter(clave_empleado=clave_empleado).exists():
                messages.error(request, 'La clave de empleado ya existe.')
            else:
                nuevo_empleado = Empleado.objects.create(
                    nombre=nombre,
                    clave_empleado=clave_empleado,
                    puesto=puesto
                )
                
                # Registrar en bitácora
                tipo_puesto = 'Administrador' if int(puesto) == 1 else 'Usuario normal'
                registrar_accion(empleado, 'Crear empleado', f'Creó el empleado {nombre} (Clave: {clave_empleado}) como {tipo_puesto}', request)
                
                messages.success(request, 'Empleado creado exitosamente.')
                return redirect('crear_empleado')
        else:
            messages.error(request, 'Todos los campos son obligatorios.')

    empleados = Empleado.objects.all()
    return render(request, 'crear_empleado.html', {'empleados': empleados})

def ver_expediente(request, id_expediente):
    if not request.session.get('empleado_id'):
        return redirect('iniciar_sesion')
    
    try:
        expediente = ConciliacionExpedientes.objects.get(id_expediente=id_expediente)
        return render(request, 'ver_expediente.html', {'expediente': expediente})
    except ConciliacionExpedientes.DoesNotExist:
        messages.error(request, 'El expediente no existe.')
        return redirect('bienvenida')

def editar_expediente(request, id_expediente):
    if not request.session.get('empleado_id'):
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=request.session['empleado_id'])
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para realizar esta acción.')
        return redirect('bienvenida')

    try:
        expediente = get_object_or_404(ConciliacionExpedientes, id_expediente=id_expediente)
        
        if request.method == 'POST':
            # Update all fields
            expediente.expediente = request.POST.get('expediente')
            expediente.junta = request.POST.get('junta', None)
            expediente.tomo = request.POST.get('tomo', None)
            expediente.actor_nombre = request.POST.get('actor_nombre', None)
            expediente.mujer_numero = request.POST.get('mujer_numero', None)
            expediente.hombre_numero = request.POST.get('hombre_numero', None)
            expediente.demandado_nombre = request.POST.get('demandado_nombre', None)
            expediente.mujer_numero_0 = request.POST.get('mujer_numero_0', None)
            expediente.hombre_numero_0 = request.POST.get('hombre_numero_0', None)
            expediente.persona_moral_numero = request.POST.get('persona_moral_numero', None)
            expediente.iebem = request.POST.get('iebem', None)
            expediente.servicios_salud = request.POST.get('servicios_salud', None)
            expediente.poder_ejecutivo = request.POST.get('poder_ejecutivo', None)
            expediente.ayuntamientos = request.POST.get('ayuntamientos', None)
            expediente.otros_organismos = request.POST.get('otros_organismos', None)
            expediente.no_se_ha_notificado_a_las_partes = request.POST.get('no_se_ha_notificado_a_las_partes', None)
            expediente.empl_no_realizado = request.POST.get('empl_no_realizado', None)
            expediente.empl_exhorto = request.POST.get('empl_exhorto', None)
            expediente.exhortos_sin_enviar = request.POST.get('exhortos_sin_enviar', None)
            expediente.exhortos_cdmx = request.POST.get('exhortos_cdmx', None)
            expediente.exhortos_foraneos = request.POST.get('exhortos_foraneos', None)
            expediente.cde = request.POST.get('cde', None)
            expediente.tercero_llamado_a_juicio = request.POST.get('tercero_llamado_a_juicio', None)
            expediente.oap = request.POST.get('oap', None)
            expediente.desahogo_pruebas = request.POST.get('desahogo_pruebas', None)
            expediente.test_falta_citar = request.POST.get('test_falta_citar', None)
            expediente.periciales_partes = request.POST.get('periciales_partes', None)
            expediente.pericial_tercero = request.POST.get('pericial_tercero', None)
            expediente.inf_falta_hacer = request.POST.get('inf_falta_hacer', None)
            expediente.inf_falta_desahogar = request.POST.get('inf_falta_desahogar', None)
            expediente.otras_pruebas = request.POST.get('otras_pruebas', None)
            expediente.alegatos = request.POST.get('alegatos', None)
            expediente.prueba_pendiente = request.POST.get('prueba_pendiente', None)
            expediente.cierre = request.POST.get('cierre', None)
            expediente.laudo_dictado = request.POST.get('laudo_dictado', None)
            expediente.absolutorio = request.POST.get('absolutorio', None)
            expediente.condenatorio = request.POST.get('condenatorio', None)
            expediente.monto = request.POST.get('monto', None)
            expediente.auto_ejecucion = request.POST.get('auto_ejecucion', None)
            expediente.terceria = request.POST.get('terceria', None)
            expediente.recurso_revision = request.POST.get('recurso_revision', None)
            expediente.remate = request.POST.get('remate', None)
            expediente.inactividad_1_anio = request.POST.get('inactividad_1_anio', None)
            expediente.inactividad_2_anios_mas = request.POST.get('inactividad_2_anios_mas', None)
            expediente.amparo_indirecto = request.POST.get('amparo_indirecto', None)
            expediente.amparo_directo_cumpl = request.POST.get('amparo_directo_cumpl', None)
            expediente.sustitucion_patronal = request.POST.get('sustitucion_patronal', None)
            expediente.no_interpuesta = request.POST.get('no_interpuesta', None)
            expediente.prescripcion = request.POST.get('prescripcion', None)
            expediente.depuracion = request.POST.get('depuracion', None)
            expediente.regularizar = request.POST.get('regularizar', None)
            expediente.reviso_capturo = request.POST.get('reviso_capturo', None)
            
            expediente.save()
            
            # Debug: verificar qué se guardó
            expediente.refresh_from_db()
            print("=== VALORES DESPUÉS DE GUARDAR ===")
            print(f"iebem: {expediente.iebem}")
            print(f"servicios_salud: {expediente.servicios_salud}")
            print(f"empl_no_realizado: {expediente.empl_no_realizado}")
            print(f"cde: {expediente.cde}")
            print(f"cierre: {expediente.cierre}")
            print("==================================")
            
            # Registrar en bitácora
            registrar_accion(empleado, 'Editar expediente', f'Editó el expediente {expediente.expediente} - Actor: {expediente.actor_nombre} vs Demandado: {expediente.demandado_nombre}', request)
            
            messages.success(request, 'Expediente actualizado correctamente.')
            return redirect('bienvenida')
        
        return render(request, 'editar_expediente.html', {'expediente': expediente})
        
    except ConciliacionExpedientes.DoesNotExist:
        messages.error(request, 'El expediente no existe.')
        return redirect('bienvenida')


def crear_empleado(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('bienvenida')

    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        clave_empleado = request.POST.get('clave_empleado')
        puesto = request.POST.get('puesto')
        if nombre and clave_empleado and puesto:
            if Empleado.objects.filter(clave_empleado=clave_empleado).exists():
                messages.error(request, 'La clave de empleado ya existe.')
            else:
                nuevo_empleado = Empleado.objects.create(
                    nombre=nombre,
                    clave_empleado=clave_empleado,
                    puesto=puesto
                )
                
                # Registrar en bitácora
                tipo_puesto = 'Administrador' if int(puesto) == 1 else 'Usuario normal'
                registrar_accion(empleado, 'Crear empleado', f'Creó el empleado {nombre} (Clave: {clave_empleado}) como {tipo_puesto}', request)
                
                messages.success(request, 'Empleado creado exitosamente.')
                return redirect('crear_empleado')
        else:
            messages.error(request, 'Todos los campos son obligatorios.')

    empleados = Empleado.objects.all()
    return render(request, 'crear_empleado.html', {'empleados': empleados})


def eliminar_expediente(request, id_expediente):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permiso para eliminar expedientes.')
        return redirect('bienvenida')
    
    try:
        expediente = ConciliacionExpedientes.objects.get(id_expediente=id_expediente)
        expediente.delete()
        messages.success(request, 'Expediente eliminado correctamente.')
    except ConciliacionExpedientes.DoesNotExist:
        messages.error(request, 'El expediente no existe.')
    
    return redirect('bienvenida')


def editar_empleado(request, id):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    admin = Empleado.objects.get(id=empleado_id)
    if not admin.es_administrador():
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('bienvenida')

    empleado_edit = Empleado.objects.get(id=id)
    if request.method == 'POST':
        empleado_edit.nombre = request.POST.get('nombre')
        empleado_edit.clave_empleado = request.POST.get('clave_empleado')
        empleado_edit.puesto = request.POST.get('puesto')
        empleado_edit.save()
        
        # Registrar en bitácora
        tipo_puesto = 'Administrador' if int(empleado_edit.puesto) == 1 else 'Usuario normal'
        registrar_accion(admin, 'Editar empleado', f'Editó el empleado {empleado_edit.nombre} (Clave: {empleado_edit.clave_empleado}) - Puesto: {tipo_puesto}', request)
        
        messages.success(request, 'Empleado actualizado correctamente.')
        return redirect('crear_empleado')
    return render(request, 'editar_empleado.html', {'empleado': empleado_edit})

def eliminar_empleado(request, id):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    admin = Empleado.objects.get(id=empleado_id)
    if not admin.es_administrador():
        messages.error(request, 'No tienes permiso para acceder a esta página.')
        return redirect('bienvenida')

    # Prevent self-deletion
    if int(id) == int(empleado_id):
        messages.error(request, 'No puedes eliminar tu propio usuario mientras estás conectado.')
        return redirect('crear_empleado')

    empleado_eliminar = Empleado.objects.get(id=id)
    nombre_eliminado = empleado_eliminar.nombre
    clave_eliminada = empleado_eliminar.clave_empleado
    
    empleado_eliminar.delete()
    
    # Registrar en bitácora
    registrar_accion(admin, 'Eliminar empleado', f'Eliminó el empleado {nombre_eliminado} (Clave: {clave_eliminada})', request)
    
    messages.success(request, 'Empleado eliminado correctamente.')
    return redirect('crear_empleado')


def cargar_expediente(request):
    if request.method == 'POST':
        empleado_id = request.session.get('empleado_id')
        expediente_id = request.POST.get('expediente_id')
        nombre_carga = request.POST.get('nombre_carga')

        try:
            empleado = Empleado.objects.get(id=empleado_id)
            expediente = DosMilSiete.objects.get(id_expediente=expediente_id)
            
            carga = CargaDescarga(
                empleado=empleado,
                expediente=expediente,
                nombre_carga=nombre_carga
            )
            carga.save()
            
            # Registrar en bitácora
            registrar_accion(empleado, 'Cargar expediente', f'Cargó el expediente {expediente.expediente} - {nombre_carga}', request)
            
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def ver_cargas(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=empleado_id)
    
    cargas = CargaDescarga.objects.all().order_by('-fecha')
    return render(request, 'ver_cargas.html', {'cargas': cargas})


# Remove @login_required decorator and keep the function as is
def archivar_expediente(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return JsonResponse({'success': False, 'error': 'Sesión no iniciada'})
    
    empleado = Empleado.objects.get(id=empleado_id)

    if request.method == 'POST':
        try:
            expedientes_ids = request.POST.getlist('expedientes_ids[]')
            motivo = request.POST.get('motivo')
            
            if not expedientes_ids:
                return JsonResponse({'success': False, 'error': 'No se seleccionaron expedientes'})
            
            archivados_count = 0
            for expediente_id in expedientes_ids:
                try:
                    expediente = DosMilSiete.objects.get(id_expediente=expediente_id)
                    
                    # Create archive record with all fields as strings
                    Archivados.objects.create(
                        expediente=expediente.expediente,
                        junta=expediente.junta,
                        actor=expediente.actor,
                        demandado=expediente.demandado,
                        motivo=motivo,
                        fecha_archivo=timezone.now()
                    )
                    
                    # Delete from original table
                    expediente.delete()
                    archivados_count += 1
                except DosMilSiete.DoesNotExist:
                    continue
            
            # Registrar en bitácora
            registrar_accion(empleado, 'Archivar expedientes', f'Archivó {archivados_count} expediente(s) - Motivo: {motivo}', request)
            
            return JsonResponse({
                'success': True, 
                'message': f'Se archivaron {archivados_count} expedientes correctamente'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def obtener_expedientes_ajax(request):
    """Vista para obtener expedientes mediante AJAX para el modal de archivado"""
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return JsonResponse({'success': False, 'error': 'Sesión no iniciada'})
    
    empleado = Empleado.objects.get(id=empleado_id)
    
    expedientes = DosMilSiete.objects.all().values('id_expediente', 'expediente', 'junta', 'actor', 'demandado')
    expedientes_list = list(expedientes)
    
    return JsonResponse({'success': True, 'expedientes': expedientes_list})

# Remove @login_required decorator and keep the function as is
def ver_archivados(request):
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=empleado_id)
    
    archivados = Archivados.objects.all().order_by('-fecha_archivo')
    return render(request, 'ver_archivados.html', {'archivados': archivados, 'empleado': empleado})

def exportar_expedientes_excel(request):
    # Verificar sesión
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        messages.error(request, 'Debes iniciar sesión')
        return redirect('iniciar_sesion')
    
    # Obtener todos los expedientes de la tabla 'expedientes'
    expedientes = ConciliacionExpedientes.objects.all()
    
    # Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Expedientes"
    
    # Estilos para el encabezado
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Definir las columnas (según los campos del modelo ConciliacionExpedientes)
    headers = [
        'EXPEDIENTE', 'JUNTA', 'TOMO', 'ACTOR', 'MUJER', 'HOMBRE',
        'DEMANDADO', 'MUJER', 'HOMBRE', 'PERSONA_MORAL',
        'IEBEM', 'SERVICIOS_SALUD', 'PODER_EJECUTIVO', 'AYUNTAMIENTOS', 'OTROS_ORGANISMOS',
        'NO_SE_HA_NOTIFICADO_PARTES', 'EMPL_NO_REALIZADO', 'EMPL_EXHORTO',
        'EXHORTOS_SIN_ENVIAR', 'EXHORTOS_CDMX', 'EXHORTOS_FORANEOS', 'C_D_E',
        'TERCERO_LLAMADO_JUICIO', 'O_A_P', 'DESAHOGO_PRUEBAS', 'TESTIMONIAL_FALTA_CITAR',
        'PERICIALES_PARTES', 'PERICIALES', 'INFORME_FALTA_HACER', 'INFORME_FALTA_DESAHOGAR',
        'OTRAS_PRUEBAS', 'ALEGATOS', 'PRUEBA_PENDIENTE', 'CIERRE', 'LAUDO_DICTADO',
        'ABSOLUTORIO', 'CONDENATORIO', 'MONTO_CONDENA', 'AUTO_EJECUCCION', 'TERCERIA',
        'RECURSO_REVISION', 'REMATE', 'INACTIVIDAD_UN_ANIO', 'INACTIVIDAD_MAS_DOS_ANIOS',
        'AMPARO_INDIRECTO', 'AMP_DIRECTO_CUMPL', 'SUSTITUCION_PATRONAL', 'NO_INTERPUESTA',
        'PRESCRIPCION', 'DEPURACION', 'REGULARIZAR', 'REVISO_CAPTURO', 'ID_EXPEDIENTE'
    ]
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[cell.column_letter].width = 18
    
    # Escribir los datos de los expedientes
    for row_num, exp in enumerate(expedientes, 2):
        ws.cell(row=row_num, column=1).value = exp.expediente or ''
        ws.cell(row=row_num, column=2).value = exp.junta or ''
        ws.cell(row=row_num, column=3).value = exp.tomo or ''
        ws.cell(row=row_num, column=4).value = exp.actor_nombre or ''
        ws.cell(row=row_num, column=5).value = exp.mujer_numero or ''
        ws.cell(row=row_num, column=6).value = exp.hombre_numero or ''
        ws.cell(row=row_num, column=7).value = exp.demandado_nombre or ''
        ws.cell(row=row_num, column=8).value = exp.mujer_numero_0 or ''
        ws.cell(row=row_num, column=9).value = exp.hombre_numero_0 or ''
        ws.cell(row=row_num, column=10).value = exp.persona_moral_numero or ''
        ws.cell(row=row_num, column=11).value = exp.iebem or ''
        ws.cell(row=row_num, column=12).value = exp.servicios_salud or ''
        ws.cell(row=row_num, column=13).value = exp.poder_ejecutivo or ''
        ws.cell(row=row_num, column=14).value = exp.ayuntamientos or ''
        ws.cell(row=row_num, column=15).value = exp.otros_organismos or ''
        ws.cell(row=row_num, column=16).value = exp.no_se_ha_notificado_a_las_partes or ''
        ws.cell(row=row_num, column=17).value = exp.empl_no_realizado or ''
        ws.cell(row=row_num, column=18).value = exp.empl_exhorto or ''
        ws.cell(row=row_num, column=19).value = exp.exhortos_sin_enviar or ''
        ws.cell(row=row_num, column=20).value = exp.exhortos_cdmx or ''
        ws.cell(row=row_num, column=21).value = exp.exhortos_foraneos or ''
        ws.cell(row=row_num, column=22).value = exp.cde or ''
        ws.cell(row=row_num, column=23).value = exp.tercero_llamado_a_juicio or ''
        ws.cell(row=row_num, column=24).value = exp.oap or ''
        ws.cell(row=row_num, column=25).value = exp.desahogo_pruebas or ''
        ws.cell(row=row_num, column=26).value = exp.test_falta_citar or ''
        ws.cell(row=row_num, column=27).value = exp.periciales_partes or ''
        ws.cell(row=row_num, column=28).value = exp.pericial_tercero or ''
        ws.cell(row=row_num, column=29).value = exp.inf_falta_hacer or ''
        ws.cell(row=row_num, column=30).value = exp.inf_falta_desahogar or ''
        ws.cell(row=row_num, column=31).value = exp.otras_pruebas or ''
        ws.cell(row=row_num, column=32).value = exp.alegatos or ''
        ws.cell(row=row_num, column=33).value = exp.prueba_pendiente or ''
        ws.cell(row=row_num, column=34).value = exp.cierre or ''
        ws.cell(row=row_num, column=35).value = exp.laudo_dictado or ''
        ws.cell(row=row_num, column=36).value = exp.absolutorio or ''
        ws.cell(row=row_num, column=37).value = exp.condenatorio or ''
        ws.cell(row=row_num, column=38).value = exp.monto or ''
        ws.cell(row=row_num, column=39).value = exp.auto_ejecucion or ''
        ws.cell(row=row_num, column=40).value = exp.terceria or ''
        ws.cell(row=row_num, column=41).value = exp.recurso_revision or ''
        ws.cell(row=row_num, column=42).value = exp.remate or ''
        ws.cell(row=row_num, column=43).value = exp.inactividad_1_anio or ''
        ws.cell(row=row_num, column=44).value = exp.inactividad_2_anios_mas or ''
        ws.cell(row=row_num, column=45).value = exp.amparo_indirecto or ''
        ws.cell(row=row_num, column=46).value = exp.amparo_directo_cumpl or ''
        ws.cell(row=row_num, column=47).value = exp.sustitucion_patronal or ''
        ws.cell(row=row_num, column=48).value = exp.no_interpuesta or ''
        ws.cell(row=row_num, column=49).value = exp.prescripcion or ''
        ws.cell(row=row_num, column=50).value = exp.depuracion or ''
        ws.cell(row=row_num, column=51).value = exp.regularizar or ''
        ws.cell(row=row_num, column=52).value = exp.reviso_capturo or ''
        ws.cell(row=row_num, column=53).value = exp.id_expediente
    
    # Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename=expedientes_{fecha_actual}.xlsx'
    
    # Guardar el libro en la respuesta
    wb.save(response)
    
    return response


def ver_bitacora(request):
    """Vista para mostrar la bitácora de acciones - solo para administradores"""
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return redirect('iniciar_sesion')
    
    empleado = Empleado.objects.get(id=empleado_id)
    if not empleado.es_administrador():
        messages.error(request, 'No tienes permisos para ver la bitácora.')
        return redirect('bienvenida')
    
    # Obtener parámetros de filtro
    filtro_empleado = request.GET.get('empleado', '')
    filtro_accion = request.GET.get('accion', '')
    filtro_fecha = request.GET.get('fecha', '')
    
    # Construir query
    bitacora = Bitacora.objects.all()
    
    if filtro_empleado:
        bitacora = bitacora.filter(empleado__nombre__icontains=filtro_empleado)
    
    if filtro_accion:
        bitacora = bitacora.filter(accion__icontains=filtro_accion)
    
    if filtro_fecha:
        bitacora = bitacora.filter(fecha_hora__date=filtro_fecha)
    
    bitacora = bitacora.order_by('-fecha_hora')[:200]  # Limitar a 200 registros recientes
    
    return render(request, 'ver_bitacora.html', {
        'bitacora': bitacora,
        'empleado': empleado,
        'filtro_empleado': filtro_empleado,
        'filtro_accion': filtro_accion,
        'filtro_fecha': filtro_fecha
    })


def restaurar_expediente(request):
    """Vista para restaurar un expediente archivado"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return JsonResponse({'success': False, 'error': 'Sesión no iniciada'})
    
    try:
        empleado = Empleado.objects.get(id=empleado_id)
        id_archivo = request.POST.get('id_archivo')
        
        if not id_archivo:
            return JsonResponse({'success': False, 'error': 'ID de archivo no proporcionado'})
        
        # Obtener el expediente archivado
        archivo = Archivados.objects.get(id_archivo=id_archivo)
        
        # Obtener el último id_expediente para generar uno nuevo
        last_expediente = DosMilSiete.objects.order_by('-id_expediente').first()
        new_id = 1 if not last_expediente else last_expediente.id_expediente + 1
        
        # Crear el expediente en DosMilSiete con los datos del archivo
        expediente_restaurado = DosMilSiete.objects.create(
            id_expediente=new_id,
            expediente=archivo.expediente,
            junta=archivo.junta,
            actor=archivo.actor,
            demandado=archivo.demandado,
            no_se_ha_notificao_a_las_partes='',
            empl_no_realizado='',
            empl_exhorto='',
            cde='',
            oap='',
            desahogo_pruebas='',
            cierre='',
            laudo_dictado='',
            auto_ejecucion='',
            prescripcion='',
            regularizar=''
        )
        
        # Registrar en bitácora
        registrar_accion(
            empleado, 
            'Restaurar expediente', 
            f'Restauró el expediente {archivo.expediente} - Actor: {archivo.actor} vs Demandado: {archivo.demandado}',
            request
        )
        
        # Eliminar el registro de archivados
        archivo.delete()
        
        return JsonResponse({
            'success': True, 
            'message': f'Expediente {archivo.expediente} restaurado correctamente'
        })
        
    except Archivados.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Expediente archivado no encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def eliminar_permanente(request):
    """Vista para eliminar permanentemente expedientes archivados - Solo administradores"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'})
    
    empleado_id = request.session.get('empleado_id')
    if not empleado_id:
        return JsonResponse({'success': False, 'error': 'Sesión no iniciada'})
    
    try:
        empleado = Empleado.objects.get(id=empleado_id)
        
        # Verificar que sea administrador
        if not empleado.es_administrador():
            return JsonResponse({'success': False, 'error': 'No tienes permisos para realizar esta acción'})
        
        import json
        archivos_ids = json.loads(request.POST.get('archivos_ids', '[]'))
        
        if not archivos_ids:
            return JsonResponse({'success': False, 'error': 'No se seleccionaron expedientes'})
        
        eliminados_count = 0
        expedientes_eliminados = []
        
        for id_archivo in archivos_ids:
            try:
                archivo = Archivados.objects.get(id_archivo=id_archivo)
                expedientes_eliminados.append(f"{archivo.expediente} - {archivo.actor} vs {archivo.demandado}")
                archivo.delete()
                eliminados_count += 1
            except Archivados.DoesNotExist:
                continue
        
        # Registrar en bitácora
        descripcion = f'Eliminó permanentemente {eliminados_count} expediente(s) archivado(s): {", ".join(expedientes_eliminados[:3])}'
        if len(expedientes_eliminados) > 3:
            descripcion += f' y {len(expedientes_eliminados) - 3} más'
        
        registrar_accion(empleado, 'Eliminar permanentemente', descripcion, request)
        
        return JsonResponse({
            'success': True,
            'message': f'Se eliminaron permanentemente {eliminados_count} expediente(s)'
        })
        
    except Empleado.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Usuario no encontrado'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
